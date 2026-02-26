import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import csv, re, io
from collections import defaultdict
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="TTS Amazon Lift Model", page_icon="📊", layout="wide")

# ═══════════════ THEME ═══════════════
BG='#0A0A0A';S1='#111111';S2='#1A1A1A';BD='#2A2A2A'
CORAL='#FF6B35';GRN='#34D399';YEL='#FBBF24';BLU='#60A5FA'
PUR='#A78BFA';T1='#FFFFFF';T2='#9CA3AF';T3='#4B5563'
CC={'HIGH':GRN,'MED':YEL,'LOW':'#FB923C','WEAK':'#EF4444','INSUF':T3}
MO=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

st.markdown(f"""<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
.stApp{{background:{BG};}}
section[data-testid="stSidebar"]{{background:{S1};border-right:1px solid {BD};}}
h1,h2,h3{{font-family:'Inter',sans-serif!important;color:{T1}!important;font-weight:800!important;}}
.kpi{{background:{S1};border:1px solid {BD};border-radius:12px;padding:18px;text-align:center;position:relative;}}
.kpi .lb{{font:700 10px 'Inter',sans-serif;color:{T2};text-transform:uppercase;letter-spacing:.08em;}}
.kpi .vl{{font:800 28px 'Inter',sans-serif;color:{T1};letter-spacing:-.02em;margin:4px 0;}}
.kpi .vg{{font:800 28px 'Inter',sans-serif;color:{GRN};letter-spacing:-.02em;margin:4px 0;}}
.kpi .sb{{font:400 10px 'Inter',sans-serif;color:{T3};}}
.kpi .tip{{display:none;position:absolute;bottom:calc(100% + 8px);left:50%;transform:translateX(-50%);background:#1F2937;color:#D1D5DB;padding:8px 12px;border-radius:8px;font:400 11px 'Inter',sans-serif;white-space:normal;width:220px;text-align:left;z-index:99;box-shadow:0 4px 12px rgba(0,0,0,.4);line-height:1.4;}}
.kpi .tip::after{{content:'';position:absolute;top:100%;left:50%;transform:translateX(-50%);border:6px solid transparent;border-top-color:#1F2937;}}
.kpi:hover .tip{{display:block;}}
.stTabs [data-baseweb="tab-list"]{{gap:4px;}}
.stTabs [data-baseweb="tab"]{{background:transparent;border:1px solid {BD};border-radius:8px;padding:6px 16px;font-size:12px;font-weight:600;color:{T2};font-family:'Inter',sans-serif;}}
.stTabs [aria-selected="true"]{{background:rgba(255,107,53,.08)!important;border-color:{CORAL}!important;color:{CORAL}!important;}}
</style>""", unsafe_allow_html=True)

# ═══════════════ HELPERS ═══════════════
def fd(v):
    if v is None: return "$0"
    v=float(v)
    if abs(v)>=1e6:return f"${v/1e6:.1f}M"
    if abs(v)>=1e3:return f"${v/1e3:.0f}K"
    return f"${v:,.0f}"
def fn(v):
    if v is None: return "0"
    v=float(v)
    if v>=1e6:return f"{v/1e6:.1f}M"
    if v>=1e3:return f"{v/1e3:.0f}K"
    return f"{v:,.0f}"
def sf(v):
    try:return float(str(v).replace('$','').replace(',','').replace('%',''))
    except:return 0.0
def kpi_h(lb,vl,sb="",g=False,tip=""):
    vc="vg" if g else "vl"
    tip_html = f'<div class="tip">{tip}</div>' if tip else ''
    return f'<div class="kpi">{tip_html}<div class="lb">{lb}</div><div class="{vc}">{vl}</div><div class="sb">{sb}</div></div>'
def sec(t):
    st.markdown(f'<div style="display:flex;align-items:center;gap:8px;margin:28px 0 10px;"><div style="width:3px;height:16px;background:{CORAL};border-radius:2px;"></div><span style="font:700 10px \'Inter\',sans-serif;color:{CORAL};text-transform:uppercase;letter-spacing:.12em;">{t}</span></div>',unsafe_allow_html=True)
def badge_h(c):
    return f'<span style="display:inline-block;font:700 9px \'Inter\',sans-serif;padding:2px 8px;border-radius:4px;letter-spacing:.06em;background:{CC.get(c,T3)}15;color:{CC.get(c,T3)};">{c}</span>'
def pthem(fig,h=350):
    fig.update_layout(height=h,plot_bgcolor=BG,paper_bgcolor=BG,font=dict(family='Inter,sans-serif',color=T2,size=10),margin=dict(l=10,r=10,t=20,b=10),hovermode='x unified',legend=dict(orientation='h',y=-0.12,font=dict(size=10)))
    fig.update_xaxes(gridcolor=BD,showline=False);fig.update_yaxes(gridcolor=BD,showline=False)
    return fig


# ═══════════════ BRAND MAP ═══════════════
BRAND_MAP={
    'Thorne Health Shop':'Thorne Research','Pure Encapsulations Shop':'Pure Encapsulations',
    'Hims & Hers':'Hims & Hers','Vital Proteins Shop':'Vital Proteins',
    'youtheory':'YouTheory','Philips Shop US':'Philips','PHILIPS':'Philips',
    'TruNiagen':'Tru Niagen','SmartMouth':'SmartMouth','Sakura of America Shop':'Sakura',
    'Strider Bikes':'Strider Bikes','Mercola Market Shop':'Dr. Mercola',
    'Natural Factors':'Natural Factors','Herbs, Etc.':'Herbs, Etc.',
    'Amazing Grass':'Amazing Grass','Emerald Labs':'Emerald Labs',
    'Balance of Nature Shop':'Balance of Nature','AdvoCare':'AdvoCare',
    'Brownmed':'Brownmed','New Chapter Inc':'New Chapter',
    'Optimum Nutrition Shop':'Optimum Nutrition','Gaia Herbs':'Gaia',
    'Atrium - Pure Encapsulations':'Pure Encapsulations','Dr Mercola':'Dr. Mercola',
    'Emerald Laboratories':'Emerald Labs','Herbs Etc.':'Herbs, Etc.',
    'Glanbia Performance Nutrition':'Optimum Nutrition',
    'Philips Avent':'Philips','Philips Norelco':'Philips','Philips Sonicare':'Philips',
    'Strider':'Strider Bikes','Youtheory':'YouTheory',
    'Advocare':'AdvoCare','Tru Niagen':'Tru Niagen',
    'Vital Proteins':'Vital Proteins','Sakura':'Sakura',
    'Thorne Research':'Thorne Research','Pure Encapsulations':'Pure Encapsulations',
    'YouTheory':'YouTheory','SmartMouth':'SmartMouth','Strider Bikes':'Strider Bikes',
    'Dr. Mercola':'Dr. Mercola','Natural Factors':'Natural Factors',
    'Herbs, Etc.':'Herbs, Etc.','Emerald Labs':'Emerald Labs',
    'Balance of Nature':'Balance of Nature','AdvoCare':'AdvoCare',
    'Brownmed':'Brownmed','New Chapter':'New Chapter',
    'Optimum Nutrition':'Optimum Nutrition','Gaia':'Gaia',
    'Amazing Grass':'Amazing Grass','Philips':'Philips',
}
def norm(name, bm):
    if not name or not str(name).strip(): return None
    name = str(name).strip()
    if name in bm: return bm[name]
    for k, v in bm.items():
        if k.lower() == name.lower(): return v
    clean = re.sub(r'\s*\(.*?\)\s*$', '', name)
    clean = re.sub(r'\s*(Shop|Official|Store|US|USA)\s*$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s*(DO NOT|DONT|no more|No Longer).*$', '', clean, flags=re.IGNORECASE).strip()
    if not clean: clean = name
    bm[name] = clean
    return clean

MO_MAP={'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
        'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
        'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}


# ═══════════════ 2024 BASELINE (pre-TTS Amazon data) ═══════════════
# Monthly organic sales, page views, ad spend, units for Jan-Dec 2024
BASELINE_2024 = {
    'AdvoCare': {'organic': [514438.83, 567508.45, 593442.22, 496137.12, 515082.92, 461819.87, 516927.72, 473747.35, 536583.29, 471417.66, 423053.95, 413966.13], 'page_views': [79541, 85602, 97650, 137386, 152580, 132947, 123152, 79633, 72812, 83202, 95621, 86742], 'ad_spend': [56889.71, 59389.08, 58600.3, 58900.2, 59141.17, 56051.49, 69770.11, 58231.44, 58689.96, 69129.65, 79256.03, 61285.54], 'units': [21162, 21727, 23796, 23773, 24249, 22871, 26992, 22780, 23676, 24855, 28598, 22948]},
    'Balance of Nature': {'organic': [1469025.44, 1476008.68, 1735130.57, 1670531.98, 1618174.19, 1455884.94, 1344471.49, 1370574.89, 1347112.64, 1413678.64, 1448036.17, 1460203.75], 'page_views': [229397, 166793, 172684, 169383, 131058, 123959, 126943, 56274, 48568, 87923, 107898, 114895], 'ad_spend': [639876.82, 623605.87, 723099.32, 788195.58, 754680.62, 655783.3, 714659.53, 226750.11, 219882.16, 277640.08, 437294.6, 461939.6], 'units': [31454, 27461, 32092, 31507, 29708, 24995, 30496, 20805, 20027, 23621, 29716, 24765]},
    'Brownmed': {'organic': [93557.66, 83744.43, 81217.75, 79244.69, 92742.67, 76666.65, 78497.33, 70703.05, 71646.23, 78371.56, 90897.46, 109969.04], 'page_views': [80430, 67114, 62003, 57493, 67669, 50439, 56687, 51666, 47473, 53048, 61202, 58422], 'ad_spend': [15844.86, 16021.74, 14018.89, 12861.62, 12905.2, 12983.95, 15669.0, 12866.76, 12766.73, 15795.55, 18118.71, 15540.67], 'units': [6560, 5663, 5371, 5225, 5872, 4871, 5552, 4877, 4793, 5660, 6614, 7090]},
    'Emerald Labs': {'organic': [299239.85, 293062.25, 336429.43, 353238.68, 355719.89, 355119.65, 336891.48, 345405.76, 346931.16, 359594.25, 352624.69, 341657.58], 'page_views': [69459, 60506, 68978, 68136, 56416, 59283, 66545, 53229, 52993, 51544, 46778, 47806], 'ad_spend': [19803.47, 19706.15, 18844.14, 18155.24, 19559.58, 19627.32, 19997.07, 20129.11, 20165.32, 20047.3, 17729.79, 23671.98], 'units': [14494, 13980, 15426, 15138, 14852, 14295, 14930, 15293, 15330, 15493, 15113, 15084]},
    'Gaia': {'organic': [1814948.12, 1927519.44, 1908470.54, 1970930.49, 2151425.16, 1937579.51, 2257750.59, 2161489.91, 2095009.64, 2147330.48, 2319699.21, 2372702.7], 'page_views': [908443, 613798, 641210, 598003, 575445, 507525, 607740, 514717, 514553, 546274, 510517, 514084], 'ad_spend': [615751.13, 353722.05, 375771.34, 350026.99, 355729.88, 310897.62, 425311.28, 372164.48, 345274.41, 428088.8, 422148.05, 449951.58], 'units': [116530, 103476, 103766, 103817, 108899, 96065, 104606, 106286, 103446, 105162, 104752, 104231]},
    'Natural Factors': {'organic': [0, 0, 0, 0, 0, 392283.3, 898281.91, 980938.1, 955010.0, 1073022.72, 1132154.66, 1151041.58], 'page_views': [0, 0, 0, 0, 0, 91059, 202029, 206684, 196184, 218462, 207902, 208570], 'ad_spend': [0, 0, 0, 0, 0, 0, 58780.19, 78226.42, 68314.21, 113623.99, 76251.92, 73364.3], 'units': [0, 0, 0, 0, 0, 14445, 39693, 46799, 45831, 53289, 54333, 56436]},
    'Optimum Nutrition': {'organic': [320207.21, 271660.99, 344342.19, 341598.82, 400389.29, 487505.16, 601143.74, 497771.04, 333667.14, 347187.49, 331351.68, 334260.07], 'page_views': [22634, 13583, 16779, 19912, 21623, 23974, 55161, 53523, 49019, 68267, 59953, 62851], 'ad_spend': [363.86, 1626.06, 6.76, 3.32, 0, 0, 0, 0, 0, 0, 0, 0], 'units': [7955, 6659, 8688, 8532, 10165, 12422, 15041, 13855, 7956, 6506, 6553, 7275]},
    'Philips': {'organic': [0, 0, 0, 0, 0, 0, 6054.93, 88244.26, 178102.65, 162221.83, 457267.62, 397915.19], 'page_views': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], 'ad_spend': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], 'units': [0, 0, 0, 0, 0, 0, 144, 1336, 3002, 2643, 14738, 12156]},
    'Pure Encapsulations': {'organic': [16287544.21, 14572545.44, 16108720.77, 15873726.06, 16402248.99, 16479088.86, 16467883.68, 15967074.04, 17551936.39, 18681341.18, 18674809.31, 19177267.29], 'page_views': [3618988, 3302257, 3289506, 3173431, 3142462, 2758562, 3168662, 3067700, 3096150, 3264343, 3241371, 3388475], 'ad_spend': [2416733.32, 2324141.84, 2544262.26, 2565358.76, 2584407.04, 2372190.05, 2707439.82, 2803608.62, 2457703.8, 2803701.9, 2731178.22, 3611254.45], 'units': [745439, 691274, 752198, 740944, 766157, 739953, 812381, 792358, 813790, 867782, 874044, 911048]},
    'Sakura': {'organic': [800959.05, 562196.26, 607430.11, 605717.19, 583244.97, 511622.28, 666306.13, 659571.18, 583302.91, 589926.07, 883472.83, 1137321.52], 'page_views': [1122705, 920058, 893199, 852929, 910952, 814641, 943675, 709444, 615002, 679498, 1012791, 1045570], 'ad_spend': [72245.16, 67959.4, 60160.84, 52165.12, 51520.11, 61161.78, 69291.5, 91900.16, 73003.47, 87752.44, 85927.11, 88043.38], 'units': [85637, 68377, 65654, 65029, 61617, 58251, 75807, 80046, 66158, 69974, 94495, 120536]},
    'SmartMouth': {'organic': [312135.82, 300211.42, 362597.83, 359819.47, 360963.45, 366789.31, 410272.87, 314358.01, 324445.52, 397777.01, 352433.61, 303491.67], 'page_views': [95678, 93366, 95745, 96261, 92445, 80792, 102245, 87393, 82096, 91440, 100500, 97474], 'ad_spend': [27885.78, 29853.67, 28959.7, 27672.94, 28253.25, 30025.69, 39548.75, 39331.3, 30734.15, 33806.7, 50389.36, 53596.63], 'units': [22386, 22676, 26578, 27939, 27265, 28425, 34855, 29782, 27311, 31861, 31121, 28375]},
    'Strider Bikes': {'organic': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 275.47], 'page_views': [0, 0, 0, 0, 0, 748, 0, 0, 0, 0, 0, 0], 'ad_spend': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], 'units': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2]},
    'Thorne Research': {'organic': [8998977.16, 8535064.03, 9473840.93, 9465218.51, 10192502.58, 10502822.89, 11063872.01, 10603157.19, 10682269.86, 13812955.02, 13767253.68, 13661194.03], 'page_views': [3091796, 2760665, 2937635, 2784771, 2261422, 2471599, 3109541, 2197050, 2220538, 2822478, 2732928, 2755428], 'ad_spend': [1457393.88, 1428177.75, 1435657.09, 1485161.14, 1555164.13, 1551551.72, 1831902.88, 2213651.83, 2222328.17, 2694774.52, 1989871.41, 2031442.36], 'units': [579719, 545983, 575412, 580296, 584646, 570695, 601056, 583141, 581449, 692043, 688618, 710253]},
    'YouTheory': {'organic': [0, 0, 0, 0, 957.72, 89528.5, 489988.62, 530388.03, 607980.83, 709357.21, 724415.01, 596355.5], 'page_views': [0, 0, 0, 1740, 67820, 259663, 285499, 269940, 348407, 445789, 420545, 356783], 'ad_spend': [0, 0, 0, 0, 0, 1535.7, 32499.51, 44096.48, 108747.45, 130175.63, 142951.92, 122632.85], 'units': [0, 0, 0, 0, 43, 4624, 32867, 34223, 43692, 57190, 59531, 54046]},
}

GMV_CAP_MULT = 5
HALO_RATE = 0.25


# ═══════════════ PARSERS ═══════════════
@st.cache_data
def parse_gmv_csv(fb):
    text=fb.decode('utf-8-sig');reader=csv.reader(io.StringIO(text));rows=list(reader)
    hi=None
    for i,r in enumerate(rows):
        if r and str(r[0]).strip().upper()=='BRAND': hi=i;break
    if hi is None: return None
    headers=rows[hi]; month_cols={}
    for ci,h in enumerate(headers):
        hl=str(h).strip().lower()
        for mn,mv in MO_MAP.items():
            if mn in hl:
                for y in ['2026','2025','2024']:
                    if y in hl: month_cols[(int(y),mv)]=ci;break
                break
    data=[]
    for r in rows[hi+1:]:
        if not r or not r[0] or r[0].strip() in ('Total',''): continue
        brand=r[0].strip();ps=r[1].strip() if len(r)>1 else '';status=r[3].strip() if len(r)>3 else ''
        monthly={}
        for (year,month),ci in month_cols.items():
            if ci<len(r): monthly[(year,month)]=sf(r[ci])
        data.append({'brand':brand,'ps':ps,'status':status,'monthly':monthly})
    return data

@st.cache_data
def parse_broadway(fb):
    import openpyxl
    wb=openpyxl.load_workbook(BytesIO(fb),read_only=True,data_only=True)
    pr,vr,ct=[],[],[]
    if 'Partner Raw' in wb.sheetnames:
        for i,row in enumerate(wb['Partner Raw'].iter_rows(values_only=True)):
            if i==0:continue
            v=list(row)
            if not v[0]:continue
            pr.append({'shop':str(v[0]),'gmv':sf(v[1]),
                'impressions':sf(v[13]) if len(v)>13 else 0,
                'visitors':sf(v[14]) if len(v)>14 else 0,
                'affiliate_gmv':sf(v[10]) if len(v)>10 else 0,
                'month':int(sf(v[18])) if len(v)>18 and v[18] else 0,
                'year':int(sf(v[20])) if len(v)>20 and v[20] else 0})
    if 'Partner Video Raw' in wb.sheetnames:
        for i,row in enumerate(wb['Partner Video Raw'].iter_rows(values_only=True)):
            if i==0:continue
            v=list(row)
            if not v[0]:continue
            vr.append({'shop':str(v[0]),
                'videos':sf(v[10]) if len(v)>10 else 0,
                'lives':sf(v[9]) if len(v)>9 else 0,
                'month':int(sf(v[13])) if len(v)>13 and v[13] else 0,
                'year':int(sf(v[15])) if len(v)>15 and v[15] else 0})
    wb.close()
    return {'pr':pr,'vr':vr}

@st.cache_data
def parse_amazon(fb):
    import openpyxl
    wb=openpyxl.load_workbook(BytesIO(fb),read_only=True,data_only=True)
    target=None
    for s in wb.sheetnames:
        ws=wb[s]
        first=[str(c).lower() if c else '' for c in next(ws.iter_rows(max_row=1,values_only=True))]
        if any('start' in f and 'date' in f for f in first) and any('brand' in f for f in first):
            target=s;break
    if not target:wb.close();return None
    ws=wb[target];headers=None;rows=[]
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        v=list(row)
        if i==0:headers=[str(c).strip() if c else '' for c in v];continue
        if v[0]:rows.append(v)
    wb.close()
    if not headers or not rows:return None
    hl=[h.lower() for h in headers]
    def fc(kws):
        for k in kws:
            for j,h in enumerate(hl):
                if all(w in h for w in k.split()):return j
        return None
    cs=fc(['start date']);cb=fc(['brand']);ct_col=fc(['total sales $','total sales'])
    cas=fc(['ad sales','advertising sales','sponsored sales'])
    cpv=fc(['total page view','page view'])
    cad=fc(['ad spend'])
    cunits=fc(['units'])
    if cs is None or cb is None or ct_col is None:return None
    data=[]
    for v in rows:
        try:
            s=v[cs]
            if isinstance(s,str):s=datetime.strptime(s.split(' ')[0],'%Y-%m-%d')
            elif not isinstance(s,datetime):continue
            sales=sf(v[ct_col]);ad_s=sf(v[cas]) if cas is not None else 0
            data.append({'year':s.year,'month':s.month,'brand_raw':str(v[cb]).strip(),
                'sales':sales,'ad_sales':ad_s,'organic':sales-ad_s,
                'page_views':sf(v[cpv]) if cpv is not None else 0,
                'ad_spend':sf(v[cad]) if cad is not None else 0,
                'units':sf(v[cunits]) if cunits is not None else 0})
        except:continue
    return data


# ═══════════════ MODEL v6 — YoY Pre/Post with 2024 Baseline ═══════════════

def build_model(gmv_data, broadway, amazon_data, bm, report_month=None):
    # Parse Amazon 2025
    amz = defaultdict(lambda: defaultdict(lambda:{'sales':0,'ad_sales':0,'organic':0,'page_views':0,'ad_spend':0,'units':0}))
    amz_brands = set()
    if amazon_data:
        for a in amazon_data:
            brand = norm(a['brand_raw'], bm)
            if not brand: continue
            amz_brands.add(brand)
            d = amz[brand][(a['year'], a['month'])]
            d['sales'] += a['sales']; d['ad_sales'] += a['ad_sales']
            d['organic'] += a['organic']; d['page_views'] += a['page_views']
            d['ad_spend'] += a['ad_spend']; d['units'] += a['units']

    # TTS monthly from GMV CSV
    tts_monthly = defaultdict(lambda: defaultdict(float))
    tts_meta = {}
    if gmv_data:
        for row in gmv_data:
            brand = norm(row['brand'], bm)
            if not brand: continue
            tts_meta[brand] = {'ps':row['ps'],'status':row['status']}
            for (y,m),gmv in row['monthly'].items():
                tts_monthly[brand][(y,m)] += gmv

    # Content from Broadway
    content_m = defaultdict(lambda: defaultdict(lambda:{'gmv':0,'impressions':0,'visitors':0,'affiliate_gmv':0,'videos':0,'lives':0}))
    if broadway:
        for p in broadway['pr']:
            if p['year']<2024:continue
            brand=norm(p['shop'],bm)
            if not brand:continue
            d=content_m[brand][(p['year'],p['month'])]
            d['gmv']+=p['gmv'];d['impressions']+=p['impressions'];d['visitors']+=p['visitors'];d['affiliate_gmv']+=p['affiliate_gmv']
        for v in broadway['vr']:
            if v['year']<2024:continue
            brand=norm(v['shop'],bm)
            if not brand:continue
            content_m[brand][(v['year'],v['month'])]['videos']+=v['videos']
            content_m[brand][(v['year'],v['month'])]['lives']+=v['lives']

    # Detect report month
    if report_month:
        latest = report_month
    else:
        content_months = set()
        for b,ms in content_m.items():
            for k in ms: content_months.add(k)
        latest = max(content_months) if content_months else (2026,1)

    # Build per-brand
    master_brands = amz_brands if amz_brands else set(tts_monthly.keys())
    brands = []

    for brand in sorted(master_brands):
        meta = tts_meta.get(brand,{})
        tts_2025 = [tts_monthly[brand].get((2025,m),0) for m in range(1,13)]
        total_tts = sum(tts_2025)

        # 2025 Amazon data
        post_organic = [amz[brand][(2025,m)]['organic'] for m in range(1,13)]
        post_pv = [amz[brand][(2025,m)]['page_views'] for m in range(1,13)]
        post_ad_spend = [amz[brand][(2025,m)]['ad_spend'] for m in range(1,13)]
        post_units = [amz[brand][(2025,m)]['units'] for m in range(1,13)]
        post_sales = [amz[brand][(2025,m)]['sales'] for m in range(1,13)]

        # 2024 baseline
        bl = BASELINE_2024.get(brand, None)
        pre_organic = bl['organic'] if bl else [0]*12
        pre_pv = bl['page_views'] if bl else [0]*12
        pre_ad_spend = bl['ad_spend'] if bl else [0]*12
        pre_units = bl['units'] if bl else [0]*12

        has_baseline = bl is not None and sum(pre_organic) > 0
        has_2025 = sum(post_organic) > 0
        has_tts = total_tts > 1000

        # TTS launch detection
        all_tts = [tts_monthly[brand].get((2024,m),0) for m in range(1,13)] + tts_2025
        launch_idx = None
        for i, v in enumerate(all_tts):
            if v > 500: launch_idx = i; break
        launch_year = 2024 if launch_idx is not None and launch_idx < 12 else 2025
        launch_month = ((launch_idx % 12) + 1) if launch_idx is not None else 0

        # Latest month values
        lc = content_m[brand].get(latest,{})
        imp = lc.get('impressions',0) if isinstance(lc,dict) else 0
        vis = lc.get('visitors',0) if isinstance(lc,dict) else 0
        vid = lc.get('videos',0) if isinstance(lc,dict) else 0
        liv = lc.get('lives',0) if isinstance(lc,dict) else 0
        aff = lc.get('affiliate_gmv',0) if isinstance(lc,dict) else 0
        jan_tts = tts_monthly[brand].get(latest,0)
        if jan_tts == 0: jan_tts = lc.get('gmv',0) if isinstance(lc,dict) else 0
        jan_amz = amz[brand].get(latest,{}).get('sales',0)
        if jan_amz == 0:
            for m in range(12,0,-1):
                jan_amz = amz[brand].get((2025,m),{}).get('sales',0)
                if jan_amz > 0: break

        # ═══ ATTRIBUTION MODEL ═══
        attributed_lift = 0; confidence = 'INSUF'; pv_signal = 'N/A'
        unexplained = 0; ad_halo = 0; tts_share = 0
        raw_attributed = 0; gmv_cap = 0; capped = False
        inc_pv = 0; inc_units = 0
        total_org_pct = 0; pv_change_pct = 0; ad_spend_pct = 0

        if has_baseline and has_2025:
            t_pre_org = sum(pre_organic); t_post_org = sum(post_organic)
            total_org_change = t_post_org - t_pre_org
            total_org_pct = total_org_change / t_pre_org if t_pre_org > 0 else 0

            t_pre_pv = sum(pre_pv); t_post_pv = sum(post_pv)
            pv_change_pct = (t_post_pv - t_pre_pv) / t_pre_pv if t_pre_pv > 0 else 0

            t_pre_ads = sum(pre_ad_spend); t_post_ads = sum(post_ad_spend)
            ad_spend_pct = (t_post_ads - t_pre_ads) / t_pre_ads if t_pre_ads > 0 else 0

            # Ad halo
            ad_halo = max(0, ad_spend_pct * t_pre_org * HALO_RATE) if ad_spend_pct > 0 else 0
            unexplained = max(0, total_org_change - ad_halo)

            # TTS share
            tts_intensity = total_tts / t_post_org if t_post_org > 0 else 0
            if pv_change_pct > ad_spend_pct and total_org_change > 0:
                tts_share = min(0.5, tts_intensity * 100)
                pv_signal = "POSITIVE"
            elif total_org_change > 0:
                tts_share = min(0.25, tts_intensity * 50)
                pv_signal = "NEUTRAL"
            else:
                tts_share = 0
                pv_signal = "NEGATIVE"

            raw_attributed = unexplained * tts_share
            gmv_cap = total_tts * GMV_CAP_MULT
            capped = raw_attributed > gmv_cap
            attributed_lift = min(raw_attributed, gmv_cap)

            # Incremental PVs
            if t_pre_pv > 0:
                expected_pv = max(0, ad_spend_pct) * t_pre_pv
                inc_pv_raw = max(0, (t_post_pv - t_pre_pv) - expected_pv)
                inc_pv = inc_pv_raw * tts_share
                if capped and raw_attributed > 0:
                    inc_pv = inc_pv * (attributed_lift / raw_attributed)
            else:
                inc_pv = max(0, t_post_pv) * tts_share

            # Incremental units
            t_post_units = sum(post_units)
            if t_post_org > 0 and t_post_units > 0:
                org_per_unit = t_post_org / t_post_units
                inc_units = attributed_lift / org_per_unit if org_per_unit > 0 else 0

            # Confidence
            org_grew = total_org_change > 0
            pv_grew = t_post_pv > t_pre_pv
            ad_stable = abs(ad_spend_pct) < 0.20
            if has_tts and org_grew and pv_grew and ad_stable: confidence = "HIGH"
            elif has_tts and org_grew and (pv_grew or ad_stable): confidence = "MED"
            elif has_tts and org_grew: confidence = "LOW"
            elif has_tts: confidence = "WEAK"

        elif has_2025 and has_tts and not has_baseline:
            # No 2024 baseline — new brand, use simple TTS ratio
            t_post_org = sum(post_organic)
            tts_intensity = total_tts / t_post_org if t_post_org > 0 else 0
            tts_share = min(0.5, tts_intensity * 100)
            raw_attributed = t_post_org * tts_share
            gmv_cap = total_tts * GMV_CAP_MULT
            capped = raw_attributed > gmv_cap
            attributed_lift = min(raw_attributed, gmv_cap)
            confidence = "LOW"
            pv_signal = "N/A"

        lift_per_dollar = attributed_lift / total_tts if total_tts > 0 else 0

        # Monthly detail for charts
        monthly = []
        for m in range(12):
            pre_o = pre_organic[m]; post_o = post_organic[m]
            monthly.append({
                'month': m+1, 'tts': tts_2025[m],
                'pre_org': pre_o, 'post_org': post_o,
                'org_yoy': post_o - pre_o,
                'org_yoy_pct': (post_o - pre_o) / pre_o if pre_o > 0 else 0,
                'pre_pv': pre_pv[m], 'post_pv': post_pv[m],
                'pv_yoy_pct': (post_pv[m] - pre_pv[m]) / pre_pv[m] if pre_pv[m] > 0 else 0,
                'pre_ads': pre_ad_spend[m], 'post_ads': post_ad_spend[m],
                'ad_yoy_pct': (post_ad_spend[m] - pre_ad_spend[m]) / pre_ad_spend[m] if pre_ad_spend[m] > 0 else 0,
            })

        brands.append({
            'brand':brand, 'ps':meta.get('ps',''), 'status':meta.get('status',''),
            'jan_tts':jan_tts, 'jan_amz':jan_amz, 'tts_total':total_tts,
            'launch_year':launch_year, 'launch_month':launch_month,
            'confidence':confidence, 'pv_signal':pv_signal,
            # Attribution
            'attributed_lift':attributed_lift, 'raw_attributed':raw_attributed,
            'unexplained':unexplained, 'ad_halo':ad_halo,
            'tts_share':tts_share, 'gmv_cap':gmv_cap, 'capped':capped,
            'lift_per_dollar':lift_per_dollar,
            'inc_pv':inc_pv, 'inc_units':inc_units,
            # YoY metrics
            'total_org_pct':total_org_pct, 'pv_change_pct':pv_change_pct,
            'ad_spend_pct':ad_spend_pct,
            # Content
            'impressions':imp, 'visitors':vis, 'videos':vid,
            'live_streams':liv, 'affiliate_gmv':aff,
            # Series
            'tts_2025':tts_2025, 'monthly':monthly,
            'pre_organic':pre_organic, 'post_organic':post_organic,
            'pre_pv':pre_pv, 'post_pv':post_pv,
        })

    return brands, latest


# ═══════════════ APP LAYOUT ═══════════════

st.markdown(f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px;"><div style="width:7px;height:7px;border-radius:50%;background:{CORAL};box-shadow:0 0 12px rgba(255,107,53,.25);"></div><span style="font:700 10px \'Inter\',sans-serif;color:{CORAL};text-transform:uppercase;letter-spacing:.16em;">Pattern x NextWave</span></div>',unsafe_allow_html=True)
st.markdown("# TTS to Amazon Lift Model")

sec("Upload Monthly Data")
c1,c2,c3 = st.columns(3)
with c1:
    st.markdown(f'<div style="background:{S1};border:2px dashed {BD};border-radius:12px;padding:14px 16px;text-align:center;"><span style="font:800 13px \'Inter\',sans-serif;color:{T1};">Monthly GMV</span><br><span style="font-size:10px;color:{T2};">TTS history (.csv)</span></div>',unsafe_allow_html=True)
    gmv_file=st.file_uploader("gmv",type=['csv'],label_visibility="collapsed",key="gmv")
with c2:
    st.markdown(f'<div style="background:{S1};border:2px dashed {BD};border-radius:12px;padding:14px 16px;text-align:center;"><span style="font:800 13px \'Inter\',sans-serif;color:{T1};">Broadway Tool</span><br><span style="font-size:10px;color:{T2};">Content metrics (.xlsm)</span></div>',unsafe_allow_html=True)
    bw_file=st.file_uploader("bw",type=['xlsm','xlsx'],label_visibility="collapsed",key="bw")
with c3:
    st.markdown(f'<div style="background:{S1};border:2px dashed {BD};border-radius:12px;padding:14px 16px;text-align:center;"><span style="font:800 13px \'Inter\',sans-serif;color:{CORAL};">Amazon Report *</span><br><span style="font-size:10px;color:{T2};">Brand master + sales (.xlsx)</span></div>',unsafe_allow_html=True)
    amz_file=st.file_uploader("amz",type=['xlsx'],label_visibility="collapsed",key="amz")

if not amz_file:
    st.markdown("---")
    st.warning("The **Amazon Report** is required. Upload it to continue.")
    st.stop()

bm = dict(BRAND_MAP)
gmv_data = parse_gmv_csv(gmv_file.read()) if gmv_file else None
if gmv_file: gmv_file.seek(0)
broadway = parse_broadway(bw_file.read()) if bw_file else None
if bw_file: bw_file.seek(0)
amazon_data = parse_amazon(amz_file.read())
amz_file.seek(0)

if not amazon_data:
    st.error("Could not parse Amazon report.")
    st.stop()

# Month selector
avail_months = set()
if broadway:
    for p in broadway['pr']:
        if p['year'] >= 2025 and p['month'] > 0:
            avail_months.add((p['year'], p['month']))
if not avail_months and gmv_data:
    for row in gmv_data:
        for (y, m) in row['monthly']:
            if y >= 2025: avail_months.add((y, m))
if not avail_months:
    avail_months = {(2026, 1)}

sorted_months = sorted(avail_months, reverse=True)
month_options = [f"{MO[m-1]} {y}" for y, m in sorted_months]
month_tuples = sorted_months

sec("Select Reporting Month")
sel_month_label = st.selectbox("Analyze lift for:", month_options, index=0)
sel_month_idx = month_options.index(sel_month_label)
selected_month = month_tuples[sel_month_idx]

# Sidebar
with st.sidebar:
    st.markdown(f'<span style="font:700 10px \'Inter\',sans-serif;color:{CORAL};text-transform:uppercase;letter-spacing:.16em;">Model v6 — YoY Pre/Post</span>',unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("**Model Parameters**")
    st.caption(f"GMV Cap: {GMV_CAP_MULT}x TTS GMV")
    st.caption(f"Ad Halo Rate: {HALO_RATE:.0%}")
    st.caption("Baseline: 2024 Amazon (embedded)")
    st.markdown("---")
    st.caption(f"GMV CSV: {'✓' if gmv_data else '—'}")
    st.caption(f"Broadway: {'✓' if broadway else '—'}")
    st.caption(f"Amazon: {len(amazon_data)} rows")

# Build model
brands, latest = build_model(gmv_data, broadway, amazon_data, bm, report_month=selected_month)

if not brands:
    st.error("No matching brands found.")
    st.stop()

df = pd.DataFrame(brands).sort_values('attributed_lift', ascending=False)
ml = f"{MO[latest[1]-1]} {latest[0]}"

st.markdown(f'<div style="margin:10px 0;font:700 11px \'Inter\',sans-serif;color:{T2};">{len(df)} brands matched | Report: {ml} | Model v6 — YoY Pre/Post</div>',unsafe_allow_html=True)

# ═══════════════ KPIs ═══════════════
ttts=df['jan_tts'].sum(); tamz=df['jan_amz'].sum()
t_lift=df['attributed_lift'].sum()
t_pv=df['inc_pv'].sum(); t_units=df['inc_units'].sum()
capped_ct = df['capped'].sum()

cols = st.columns(6)
with cols[0]: st.markdown(kpi_h("TTS GMV",fd(ttts),ml,tip="TikTok Shop GMV for the selected month across all brands."),unsafe_allow_html=True)
with cols[1]: st.markdown(kpi_h("AMZ Sales",fd(tamz),"latest month",tip="Total Amazon sales for the selected month."),unsafe_allow_html=True)
with cols[2]: st.markdown(kpi_h("Attributed Lift",fd(t_lift),f"{t_lift/tamz*100:.2f}% of AMZ" if tamz>0 else "",g=True,tip="Amazon organic sales attributed to TTS activity. Uses YoY pre/post comparison with 2024 baseline, ad halo discount, and 5x GMV cap."),unsafe_allow_html=True)
with cols[3]: st.markdown(kpi_h("Incr. Page Views",fn(t_pv),"attributed to TTS",g=True,tip="Incremental Amazon page views above 2024 baseline, attributed to TTS at the same share rate."),unsafe_allow_html=True)
with cols[4]: st.markdown(kpi_h("Incr. Units",fn(t_units),"estimated",g=True,tip="Estimated incremental units sold based on attributed lift and current revenue per unit."),unsafe_allow_html=True)
with cols[5]: st.markdown(kpi_h("Lift / $1 TTS",f"${t_lift/ttts:.2f}" if ttts>0 else "$0","portfolio avg",tip="Dollars of Amazon organic lift per dollar of TTS GMV. Capped at 5x per brand."),unsafe_allow_html=True)


# ═══════════════ TABS ═══════════════
tabs = st.tabs(["📊 Portfolio","🔍 Brand Deep Dive","📄 Export"])

# ── TAB 1: PORTFOLIO ──
with tabs[0]:
    sec("Attribution Summary")
    tbl = df[['brand','confidence','pv_signal','jan_tts','tts_total','jan_amz',
              'total_org_pct','ad_spend_pct','attributed_lift','inc_pv','inc_units','lift_per_dollar','capped']].copy()
    tbl.columns = ['Brand','Conf','PV Signal','TTS GMV (mo)','TTS Total','AMZ Sales',
                   'Org YoY%','Ad Spend Δ%','Attributed Lift','Incr PVs','Incr Units','$/TTS','Capped']
    tbl['TTS GMV (mo)'] = tbl['TTS GMV (mo)'].apply(fd)
    tbl['TTS Total'] = tbl['TTS Total'].apply(fd)
    tbl['AMZ Sales'] = tbl['AMZ Sales'].apply(fd)
    tbl['Org YoY%'] = tbl['Org YoY%'].apply(lambda x: f"{x:+.0%}")
    tbl['Ad Spend Δ%'] = tbl['Ad Spend Δ%'].apply(lambda x: f"{x:+.0%}")
    tbl['Attributed Lift'] = tbl['Attributed Lift'].apply(fd)
    tbl['Incr PVs'] = tbl['Incr PVs'].apply(fn)
    tbl['Incr Units'] = tbl['Incr Units'].apply(lambda x: fn(x))
    tbl['$/TTS'] = tbl['$/TTS'].apply(lambda x: f"${x:.2f}")
    tbl['Capped'] = tbl['Capped'].apply(lambda x: "⚠ 5x" if x else "—")
    st.dataframe(tbl, use_container_width=True, hide_index=True)

    sec("Attributed Lift by Brand")
    pos = df[df['attributed_lift']>0].sort_values('attributed_lift')
    if len(pos) > 0:
        fig = go.Figure()
        fig.add_trace(go.Bar(
            y=pos['brand'], x=pos['attributed_lift'],
            orientation='h', marker_color=CORAL,
            text=pos['attributed_lift'].apply(fd),
            textposition='outside', textfont=dict(size=10,color=T1),
            hovertemplate='%{y}: %{x:$,.0f}<extra></extra>'
        ))
        pthem(fig, h=max(250, len(pos)*35))
        fig.update_layout(xaxis_title="Attributed Amazon Organic Lift ($)")
        st.plotly_chart(fig, use_container_width=True)

    sec("Lift per $1 TTS by Brand")
    lpd = df[df['lift_per_dollar']>0].sort_values('lift_per_dollar')
    if len(lpd) > 0:
        fig2 = go.Figure()
        colors = [GRN if not r['capped'] else YEL for _,r in lpd.iterrows()]
        fig2.add_trace(go.Bar(
            y=lpd['brand'], x=lpd['lift_per_dollar'],
            orientation='h', marker_color=colors,
            text=lpd['lift_per_dollar'].apply(lambda x: f"${x:.2f}"),
            textposition='outside', textfont=dict(size=10,color=T1),
        ))
        pthem(fig2, h=max(250, len(lpd)*35))
        fig2.update_layout(xaxis_title="$ Lift per $1 TTS GMV")
        st.plotly_chart(fig2, use_container_width=True)
        st.caption("🟢 Uncapped  🟡 Capped at 5x GMV")

    sec("YoY Organic Growth vs TTS Activity")
    has_data = df[(df['total_org_pct'] != 0) & (df['tts_total'] > 0)]
    if len(has_data) > 0:
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(
            x=has_data['tts_total'], y=has_data['total_org_pct']*100,
            mode='markers+text', text=has_data['brand'],
            textposition='top center', textfont=dict(size=9, color=T2),
            marker=dict(size=12, color=CORAL, opacity=0.8),
            hovertemplate='%{text}<br>TTS: %{x:$,.0f}<br>Org YoY: %{y:.1f}%<extra></extra>'
        ))
        fig3.add_hline(y=0, line_dash="dash", line_color=T3)
        pthem(fig3, h=400)
        fig3.update_layout(xaxis_title="Total TTS GMV (2025)", yaxis_title="Organic Sales YoY Change %",
                          xaxis_type="log")
        st.plotly_chart(fig3, use_container_width=True)

# ── TAB 2: BRAND DEEP DIVE ──
with tabs[1]:
    brand_list = sorted(df['brand'].unique())
    sel_brand = st.selectbox("Select Brand", brand_list, key="brand_picker")
    br = df[df['brand']==sel_brand].iloc[0]

    # KPI row
    c1,c2,c3,c4,c5 = st.columns(5)
    with c1: st.markdown(kpi_h("TTS GMV (mo)",fd(br['jan_tts']),ml),unsafe_allow_html=True)
    with c2: st.markdown(kpi_h("AMZ Sales",fd(br['jan_amz']),"latest"),unsafe_allow_html=True)
    with c3: st.markdown(kpi_h("Attributed Lift",fd(br['attributed_lift']),
        f"{'⚠ Capped 5x' if br['capped'] else '✓ Uncapped'}",g=True),unsafe_allow_html=True)
    with c4: st.markdown(kpi_h("Lift / $1",f"${br['lift_per_dollar']:.2f}",br['confidence']),unsafe_allow_html=True)
    with c5: st.markdown(kpi_h("Confidence",br['confidence'],br['pv_signal']),unsafe_allow_html=True)

    # Attribution waterfall
    sec("Attribution Waterfall")
    wf_labels = ['Organic Δ YoY', 'Ad Halo', 'Unexplained', f'TTS Share ({br["tts_share"]:.0%})', 'GMV Cap' if br['capped'] else 'Final']
    wf_vals = [br.get('total_org_pct',0) * sum(br['pre_organic']) if sum(br['pre_organic'])>0 else sum(br['post_organic']),
               -br['ad_halo'], br['unexplained'], -br['unexplained']*(1-br['tts_share']) if br['tts_share']<1 else 0, 0]

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"""
        <div style="background:{S1};border:1px solid {BD};border-radius:12px;padding:20px;">
        <div style="font:700 12px 'Inter',sans-serif;color:{T1};margin-bottom:12px;">Model Breakdown</div>
        <div style="font:400 11px 'Inter',sans-serif;color:{T2};line-height:2.2;">
        Organic YoY Δ: <span style="color:{T1}">{fd(sum(br['post_organic'])-sum(br['pre_organic']))}</span> ({br['total_org_pct']:+.1%})<br>
        Page Views YoY: <span style="color:{T1}">{br['pv_change_pct']:+.1%}</span> [{br['pv_signal']}]<br>
        Ad Spend YoY: <span style="color:{T1}">{br['ad_spend_pct']:+.1%}</span><br>
        <hr style="border-color:{BD};margin:6px 0;">
        Ad Halo Discount: <span style="color:{T1}">-{fd(br['ad_halo'])}</span><br>
        Unexplained Organic: <span style="color:{T1}">{fd(br['unexplained'])}</span><br>
        TTS Share: <span style="color:{T1}">{br['tts_share']:.1%}</span><br>
        Raw Attributed: <span style="color:{T1}">{fd(br['raw_attributed'])}</span><br>
        {'<span style="color:'+YEL+'">⚠ Capped at 5x: '+fd(br["gmv_cap"])+'</span><br>' if br['capped'] else ''}
        <hr style="border-color:{BD};margin:6px 0;">
        <span style="color:{GRN};font-weight:700;">Final Lift: {fd(br['attributed_lift'])}</span><br>
        Incr. Page Views: <span style="color:{T1}">{fn(br['inc_pv'])}</span><br>
        Incr. Units: <span style="color:{T1}">{fn(br['inc_units'])}</span><br>
        Lift per $1 TTS: <span style="color:{T1}">${br['lift_per_dollar']:.2f}</span>
        </div></div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div style="background:{S1};border:1px solid {BD};border-radius:12px;padding:20px;">
        <div style="font:700 12px 'Inter',sans-serif;color:{T1};margin-bottom:12px;">Content Metrics ({ml})</div>
        <div style="font:400 11px 'Inter',sans-serif;color:{T2};line-height:2.2;">
        TTS GMV: <span style="color:{T1}">{fd(br['jan_tts'])}</span><br>
        TTS Impressions: <span style="color:{T1}">{fn(br['impressions'])}</span><br>
        TTS Visitors: <span style="color:{T1}">{fn(br['visitors'])}</span><br>
        Videos: <span style="color:{T1}">{fn(br['videos'])}</span><br>
        Live Streams: <span style="color:{T1}">{fn(br['live_streams'])}</span><br>
        Affiliate GMV: <span style="color:{T1}">{fd(br['affiliate_gmv'])}</span><br>
        <hr style="border-color:{BD};margin:6px 0;">
        TTS Total (2025): <span style="color:{T1}">{fd(br['tts_total'])}</span><br>
        Launch: <span style="color:{T1}">{MO[br['launch_month']-1] if br['launch_month']>0 else 'N/A'} {br['launch_year']}</span>
        </div></div>
        """, unsafe_allow_html=True)

    # YoY chart
    sec("YoY Organic Sales Comparison (2024 vs 2025)")
    monthly = br['monthly']
    fig_yoy = make_subplots(specs=[[{"secondary_y": True}]])
    fig_yoy.add_trace(go.Bar(name='2024 Organic', x=MO, y=[m['pre_org'] for m in monthly],
        marker_color=T3, opacity=0.5), secondary_y=False)
    fig_yoy.add_trace(go.Bar(name='2025 Organic', x=MO, y=[m['post_org'] for m in monthly],
        marker_color=GRN, opacity=0.8), secondary_y=False)
    fig_yoy.add_trace(go.Scatter(name='TTS GMV', x=MO, y=[m['tts'] for m in monthly],
        mode='lines+markers', line=dict(color=CORAL, width=2), marker=dict(size=6)),
        secondary_y=True)
    pthem(fig_yoy, h=380)
    fig_yoy.update_layout(barmode='group')
    fig_yoy.update_yaxes(title_text="Organic Sales", secondary_y=False)
    fig_yoy.update_yaxes(title_text="TTS GMV", secondary_y=True)
    st.plotly_chart(fig_yoy, use_container_width=True)

    # Page views YoY
    sec("YoY Page Views (2024 vs 2025)")
    fig_pv = go.Figure()
    fig_pv.add_trace(go.Bar(name='2024 PVs', x=MO, y=[m['pre_pv'] for m in monthly],
        marker_color=T3, opacity=0.5))
    fig_pv.add_trace(go.Bar(name='2025 PVs', x=MO, y=[m['post_pv'] for m in monthly],
        marker_color=BLU, opacity=0.8))
    pthem(fig_pv, h=300)
    fig_pv.update_layout(barmode='group')
    st.plotly_chart(fig_pv, use_container_width=True)

    # Monthly detail table
    sec("Monthly Detail")
    mt = pd.DataFrame(monthly)
    mt['Mo'] = [MO[i] for i in range(12)]
    mt_disp = mt[['Mo','tts','pre_org','post_org','org_yoy','org_yoy_pct','pre_pv','post_pv','pv_yoy_pct','ad_yoy_pct']].copy()
    mt_disp.columns = ['Month','TTS GMV','2024 Organic','2025 Organic','Org Δ','Org Δ%','2024 PVs','2025 PVs','PV Δ%','Ad Spend Δ%']
    for c in ['TTS GMV','2024 Organic','2025 Organic','Org Δ']:
        mt_disp[c] = mt_disp[c].apply(fd)
    for c in ['2024 PVs','2025 PVs']:
        mt_disp[c] = mt_disp[c].apply(fn)
    for c in ['Org Δ%','PV Δ%','Ad Spend Δ%']:
        mt_disp[c] = mt_disp[c].apply(lambda x: f"{x:+.0%}" if abs(x) < 100 else f"{x:+.0f}x")
    st.dataframe(mt_disp, use_container_width=True, hide_index=True)


# ── TAB 3: EXPORT ──
with tabs[2]:
    sec("PDF Brand Report")
    pdf_brand_list = sorted(df['brand'].unique())
    pdf_brand = st.selectbox("Generate PDF for:", pdf_brand_list, key="pdf_brand_picker")

    if st.button("Generate PDF Report", type="primary"):
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        from reportlab.lib.colors import HexColor

        pbr = df[df['brand']==pdf_brand].iloc[0]
        monthly_d = pbr['monthly']
        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=letter)
        W, H = letter
        bg = HexColor('#0A0A0A'); wh = HexColor('#FFFFFF'); coral_c = HexColor('#FF6B35')
        grn_c = HexColor('#34D399'); t2c = HexColor('#9CA3AF'); bdc = HexColor('#2A2A2A')

        def draw_bg():
            c.setFillColor(bg); c.rect(0, 0, W, H, fill=1)

        def chart_to_img(fig_func, w=5.5, h=2.5):
            fig, ax = fig_func(w, h)
            fig.patch.set_facecolor('#0A0A0A')
            ax.set_facecolor('#0A0A0A')
            ax.tick_params(colors='#9CA3AF', labelsize=7)
            for spine in ax.spines.values(): spine.set_color('#2A2A2A')
            buf2 = BytesIO()
            fig.savefig(buf2, format='png', dpi=150, bbox_inches='tight', facecolor='#0A0A0A')
            plt.close(fig)
            buf2.seek(0)
            return ImageReader(buf2)

        # ═══ PAGE 1 ═══
        draw_bg()
        # Header
        c.setFillColor(coral_c); c.setFont("Helvetica-Bold", 8)
        c.drawString(40, H-35, "PATTERN x NEXTWAVE")
        c.setFillColor(wh); c.setFont("Helvetica-Bold", 20)
        c.drawString(40, H-60, f"{pdf_brand} — TTS Lift Report")
        c.setFillColor(t2c); c.setFont("Helvetica", 9)
        c.drawString(40, H-78, f"Model v6 | YoY Pre/Post | {ml} | 2024 Baseline")

        # KPI boxes
        kpis = [
            ("TTS GMV", fd(pbr['jan_tts'])),
            ("AMZ Sales", fd(pbr['jan_amz'])),
            ("Attributed Lift", fd(pbr['attributed_lift'])),
            ("Incr. PVs", fn(pbr['inc_pv'])),
            ("Incr. Units", fn(pbr['inc_units'])),
            ("Lift/$1", f"${pbr['lift_per_dollar']:.2f}"),
        ]
        bw = 82; bx = 40
        for lb, vl in kpis:
            c.setFillColor(HexColor('#111111')); c.roundRect(bx, H-130, bw, 40, 5, fill=1)
            c.setStrokeColor(bdc); c.roundRect(bx, H-130, bw, 40, 5, fill=0, stroke=1)
            c.setFillColor(t2c); c.setFont("Helvetica-Bold", 6)
            c.drawCentredString(bx+bw/2, H-100, lb.upper())
            c.setFillColor(wh); c.setFont("Helvetica-Bold", 13)
            c.drawCentredString(bx+bw/2, H-120, vl)
            bx += bw + 6

        # Confidence + PV Signal
        c.setFillColor(t2c); c.setFont("Helvetica", 8)
        c.drawString(40, H-148, f"Confidence: {pbr['confidence']}  |  PV Signal: {pbr['pv_signal']}  |  {'⚠ Capped at 5x' if pbr['capped'] else '✓ Uncapped'}")

        # YoY Organic chart
        def yoy_chart(w, h):
            fig, ax = plt.subplots(figsize=(w, h))
            x = np.arange(12)
            ax.bar(x - 0.2, [m['pre_org'] for m in monthly_d], 0.35, color='#4B5563', alpha=0.6, label='2024')
            ax.bar(x + 0.2, [m['post_org'] for m in monthly_d], 0.35, color='#34D399', alpha=0.8, label='2025')
            ax2 = ax.twinx()
            ax2.plot(x, [m['tts'] for m in monthly_d], 'o-', color='#FF6B35', linewidth=1.5, markersize=4, label='TTS GMV')
            ax2.tick_params(colors='#9CA3AF', labelsize=7)
            ax.set_xticks(x); ax.set_xticklabels(MO, fontsize=7)
            ax.legend(loc='upper left', fontsize=6, facecolor='#0A0A0A', edgecolor='#2A2A2A', labelcolor='#9CA3AF')
            ax2.legend(loc='upper right', fontsize=6, facecolor='#0A0A0A', edgecolor='#2A2A2A', labelcolor='#9CA3AF')
            ax.set_title('YoY Organic Sales + TTS GMV', color='white', fontsize=9, fontweight='bold', pad=8)
            return fig, ax

        img = chart_to_img(yoy_chart, 7, 2.5)
        c.drawImage(img, 30, H-400, width=540, height=230, mask='auto')

        # Attribution waterfall text
        y = H-420
        c.setFillColor(coral_c); c.setFont("Helvetica-Bold", 8); c.drawString(40, y, "ATTRIBUTION WATERFALL")
        items = [
            (f"Organic Δ YoY:", fd(sum(pbr['post_organic'])-sum(pbr['pre_organic'])), f"({pbr['total_org_pct']:+.1%})"),
            (f"Ad Halo Discount:", f"-{fd(pbr['ad_halo'])}", f"({HALO_RATE:.0%} of ad spend Δ)"),
            (f"Unexplained Organic:", fd(pbr['unexplained']), ""),
            (f"TTS Share:", f"{pbr['tts_share']:.1%}", f"(PV signal: {pbr['pv_signal']})"),
            (f"Raw Attributed:", fd(pbr['raw_attributed']), f"{'⚠ Capped at '+fd(pbr['gmv_cap']) if pbr['capped'] else ''}"),
            (f"FINAL ATTRIBUTED LIFT:", fd(pbr['attributed_lift']), ""),
        ]
        y -= 16
        for lb, vl, note in items:
            c.setFillColor(t2c); c.setFont("Helvetica", 8); c.drawString(55, y, lb)
            is_final = "FINAL" in lb
            c.setFillColor(grn_c if is_final else wh); c.setFont("Helvetica-Bold", 9 if is_final else 8)
            c.drawString(230, y, vl)
            if note:
                c.setFillColor(t2c); c.setFont("Helvetica", 7); c.drawString(340, y, note)
            y -= 14

        # ═══ PAGE 2 ═══
        c.showPage(); draw_bg()
        c.setFillColor(coral_c); c.setFont("Helvetica-Bold", 8); c.drawString(40, H-35, "PATTERN x NEXTWAVE")
        c.setFillColor(wh); c.setFont("Helvetica-Bold", 16); c.drawString(40, H-55, f"{pdf_brand} — Monthly Detail")

        # Page views chart
        def pv_chart(w, h):
            fig, ax = plt.subplots(figsize=(w, h))
            x = np.arange(12)
            ax.bar(x - 0.2, [m['pre_pv'] for m in monthly_d], 0.35, color='#4B5563', alpha=0.6, label='2024 PVs')
            ax.bar(x + 0.2, [m['post_pv'] for m in monthly_d], 0.35, color='#60A5FA', alpha=0.8, label='2025 PVs')
            ax.set_xticks(x); ax.set_xticklabels(MO, fontsize=7)
            ax.legend(fontsize=6, facecolor='#0A0A0A', edgecolor='#2A2A2A', labelcolor='#9CA3AF')
            ax.set_title('YoY Page Views', color='white', fontsize=9, fontweight='bold', pad=8)
            return fig, ax

        img2 = chart_to_img(pv_chart, 7, 2.2)
        c.drawImage(img2, 30, H-270, width=540, height=200, mask='auto')

        # Monthly table
        c.setFillColor(coral_c); c.setFont("Helvetica-Bold", 8); c.drawString(40, H-290, "MONTHLY COMPARISON")
        cols_x = [40, 75, 135, 210, 285, 340, 395, 455, 510]
        col_labels = ['Mo', 'TTS GMV', '2024 Org', '2025 Org', 'Org Δ%', '2024 PV', '2025 PV', 'PV Δ%', 'Ad Δ%']
        y = H-306
        c.setFillColor(t2c); c.setFont("Helvetica-Bold", 6)
        for i, lb in enumerate(col_labels):
            c.drawString(cols_x[i], y, lb)
        y -= 3
        c.setStrokeColor(bdc); c.line(40, y, 560, y)
        y -= 12
        c.setFont("Helvetica", 7)
        for md in monthly_d:
            c.setFillColor(wh)
            vals = [
                MO[md['month']-1],
                fd(md['tts']),
                fd(md['pre_org']),
                fd(md['post_org']),
                f"{md['org_yoy_pct']:+.0%}" if abs(md['org_yoy_pct']) < 100 else f"{md['org_yoy_pct']:+.0f}x",
                fn(md['pre_pv']),
                fn(md['post_pv']),
                f"{md['pv_yoy_pct']:+.0%}" if abs(md['pv_yoy_pct']) < 100 else f"{md['pv_yoy_pct']:+.0f}x",
                f"{md['ad_yoy_pct']:+.0%}" if abs(md['ad_yoy_pct']) < 100 else "N/A",
            ]
            for i, v in enumerate(vals):
                c.drawString(cols_x[i], y, v)
            y -= 12

        # Footer
        c.setFillColor(t2c); c.setFont("Helvetica", 7)
        c.drawString(40, 30, f"Pattern x NextWave | TTS → Amazon Lift Model v6 | {ml}")

        c.save()
        buf.seek(0)
        st.download_button(f"Download {pdf_brand} Report (PDF)", buf.getvalue(),
                          f"{pdf_brand.replace(' ','_')}_lift_report.pdf", "application/pdf")

    sec("CSV Export")
    ec = df[['brand','confidence','pv_signal','jan_tts','tts_total','jan_amz',
             'total_org_pct','ad_spend_pct','attributed_lift','inc_pv','inc_units',
             'lift_per_dollar','capped','tts_share','unexplained','ad_halo']].copy()
    ec.columns = ['Brand','Confidence','PV Signal','TTS GMV (mo)','TTS Total','AMZ Sales',
                  'Org YoY%','Ad Spend Δ%','Attributed Lift','Incr PVs','Incr Units',
                  '$/TTS','Capped','TTS Share','Unexplained Org','Ad Halo']
    csv_out = ec.to_csv(index=False)
    st.download_button("Download Attribution Summary (CSV)", csv_out, "tts_lift_v6.csv", "text/csv")

st.caption(f"Pattern x NextWave | TTS → Amazon Lift Model v6 | {ml} | {len(df)} brands")

